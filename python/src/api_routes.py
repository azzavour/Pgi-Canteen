import json
import math
import sqlite3
from typing import List, Optional, Literal, cast, Any
from collections import defaultdict
import asyncio
import datetime
from datetime import date
import io
import calendar
import os
import time
from uuid import uuid4
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Request, HTTPException, status, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, Response
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from . import sse_manager
from .sqlite_database import get_db_connection
from .email_service import send_order_confirmation
from .portal_auth import verify_portal_token
from .quota_utils import evaluate_tenant_quota_for_today
from .tenant_utils import generate_verification_code
from update_employee_email import update_employee_email
from .portal_control import read_override

router = APIRouter()

try:
    JAKARTA_TZ = ZoneInfo("Asia/Jakarta")
except ZoneInfoNotFoundError:
    JAKARTA_TZ = datetime.timezone(datetime.timedelta(hours=7))

CANTEEN_OPEN_HOUR = 8
CANTEEN_CLOSE_HOUR = 11

DAILY_TRANSACTION_LIMIT_MESSAGE = (
    "Anda sudah melakukan transaksi hari ini (preorder/tap). Hanya 1 transaksi per hari."
)

def ensure_dashboard_admins_table(conn: sqlite3.Connection) -> None:
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS dashboard_admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT UNIQUE NOT NULL,
            name TEXT
        )
        """
    )
    conn.commit()


def _get_local_day_bounds_from_string(date_text: Optional[str] = None) -> tuple[str, str]:
    """
    Returns (start_of_day, end_of_day) strings in local time for the provided date.
    """
    if date_text:
        normalized = date_text.strip()
        if not normalized:
            base_date = datetime.datetime.now().date()
        else:
            normalized = normalized.split("T")[0][:10]
            try:
                base_date = datetime.date.fromisoformat(normalized)
            except ValueError:
                base_date = datetime.datetime.fromisoformat(normalized).date()
    else:
        base_date = datetime.datetime.now().date()

    start_dt = datetime.datetime.combine(base_date, datetime.time.min)
    end_dt = start_dt + datetime.timedelta(days=1)
    return (
        start_dt.strftime("%Y-%m-%d %H:%M:%S"),
        end_dt.strftime("%Y-%m-%d %H:%M:%S"),
    )


def _normalize_timestamp_to_local(
    raw_ts: Optional[str], *, field_name: str = "timestamp"
) -> tuple[datetime.datetime, str, str]:
    """
    Normalizes raw timestamp string (or None) into Asia/Jakarta timezone.
    Returns tuple of (aware datetime, transaction_date_text, transaction_day).
    """
    local_tz = JAKARTA_TZ
    if not raw_ts:
        local_dt = datetime.datetime.now(tz=local_tz)
    else:
        normalized = raw_ts.strip()
        if not normalized:
            local_dt = datetime.datetime.now(tz=local_tz)
        else:
            if normalized.endswith("Z"):
                normalized = normalized[:-1] + "+00:00"
            try:
                parsed = datetime.datetime.fromisoformat(normalized)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"{field_name} harus berupa timestamp ISO8601 yang valid.",
                )
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=local_tz)
            local_dt = parsed.astimezone(local_tz)
    return (
        local_dt,
        local_dt.strftime("%Y-%m-%d %H:%M:%S"),
        local_dt.date().isoformat(),
    )


def generate_ticket_number(
    order_datetime: datetime.datetime, sequence_number: int
) -> str:
    if sequence_number >= 0:
        return f"{sequence_number:03d}"
    return str(sequence_number)


def get_tenant_prefix(tenant_name: Optional[str]) -> str:
    name = (tenant_name or "").lower()
    if "yanti" in name:
        return "A"
    if "rima" in name:
        return "B"
    return ""


def get_whatsapp_number_for_tenant(tenant_name: str) -> Optional[str]:
    mapping = {
        "yanti": "6285880259653",
        "rima": "6285718899709",
    }
    name_lower = tenant_name.lower()
    for key, number in mapping.items():
        if key in name_lower:
            return number
    return None


def is_within_operational_hours(
    open_hour: int, close_hour: int, now: Optional[datetime.datetime] = None
) -> bool:
    now = now or datetime.datetime.now()
    current_time = now.time()
    start_time = datetime.time(open_hour, 0)
    end_time = datetime.time(close_hour, 0)
    return start_time <= current_time < end_time


def _normalize_canteen_mode(raw_value: Optional[str]) -> Literal["OPEN", "CLOSE", "NORMAL"]:
    value = (raw_value or "NORMAL").strip().upper()
    if value not in {"OPEN", "CLOSE", "NORMAL"}:
        value = "NORMAL"
    return cast(Literal["OPEN", "CLOSE", "NORMAL"], value)


def _get_or_create_canteen_status_row() -> dict[str, Any]:
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT mode, updated_at, updated_by FROM canteen_status WHERE id = 1")
        row = cursor.fetchone()
        if not row:
            cursor.execute(
                """
                INSERT INTO canteen_status (id, mode, updated_at, updated_by)
                VALUES (1, 'NORMAL', datetime('now','localtime'), NULL)
                """
            )
            conn.commit()
            cursor.execute("SELECT mode, updated_at, updated_by FROM canteen_status WHERE id = 1")
            row = cursor.fetchone()
        return dict(row)
    finally:
        conn.close()


def get_canteen_mode() -> Literal["OPEN", "CLOSE", "NORMAL"]:
    row = _get_or_create_canteen_status_row()
    return _normalize_canteen_mode(row.get("mode"))


def get_effective_canteen_mode() -> Literal["OPEN", "CLOSE", "NORMAL"]:
    override = read_override()
    if override == "open":
        return "OPEN"
    if override == "closed":
        return "CLOSE"
    return get_canteen_mode()


def update_canteen_mode(new_mode: Literal["OPEN", "CLOSE", "NORMAL"], updated_by: Optional[str] = None) -> None:
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO canteen_status (id, mode, updated_at, updated_by)
            VALUES (1, ?, datetime('now','localtime'), ?)
            ON CONFLICT(id) DO UPDATE SET
                mode=excluded.mode,
                updated_at=excluded.updated_at,
                updated_by=excluded.updated_by
            """,
            (new_mode, updated_by),
        )
        conn.commit()
    finally:
        conn.close()


def canteen_is_open(now: Optional[datetime.datetime] = None) -> bool:
    now = now or datetime.datetime.now()
    mode = get_effective_canteen_mode()
    if mode == "OPEN":
        return True
    if mode == "CLOSE":
        return False
    return is_within_operational_hours(CANTEEN_OPEN_HOUR, CANTEEN_CLOSE_HOUR, now)


def get_canteen_status(now: datetime.datetime) -> dict:
    """
    Determine whether the canteen is open or closed.
    Allows manual override via portal_control file.
    """
    mode = get_effective_canteen_mode()
    open_time_text = f"{CANTEEN_OPEN_HOUR:02d}:00"
    close_time_text = f"{CANTEEN_CLOSE_HOUR:02d}:00"

    status_payload = {
        "is_open": False,
        "reason": "",
        "message": "",
        "open_time": open_time_text,
        "close_time": close_time_text,
        "mode": mode,
    }

    if mode == "OPEN":
        status_payload.update(
            {
                "is_open": True,
                "reason": "forced_open",
                "message": "Portal dibuka manual melalui panel admin.",
            }
        )
        return status_payload
    if mode == "CLOSE":
        status_payload.update(
            {
                "is_open": False,
                "reason": "forced_closed",
                "message": "Portal ditutup manual melalui panel admin.",
            }
        )
        return status_payload

    if now.time() < datetime.time(CANTEEN_OPEN_HOUR, 0):
        status_payload["reason"] = "before_open"
        status_payload["message"] = (
            f"Akan dibuka pada pukul {open_time_text} WIB, silakan kembali lagi pada waktu tersebut."
        )
    elif now.time() >= datetime.time(CANTEEN_CLOSE_HOUR, 0):
        status_payload["reason"] = "after_close"
        status_payload["message"] = (
            "Silakan pesan makan langsung di kantin (on the spot)."
        )
    else:
        status_payload["is_open"] = True
        status_payload["reason"] = "open"
        status_payload["message"] = (
            f"Cawang Canteen buka. Jam layanan pemesanan: {open_time_text}–{close_time_text} WIB."
        )

    status_payload["is_open"] = canteen_is_open(now)
    return status_payload


def _card_has_transaction_for_day(cursor, card_number: str, transaction_day: str) -> bool:
    if not card_number or not transaction_day:
        return False
    cursor.execute(
        """
        SELECT 1
        FROM transactions
        WHERE card_number = ?
          AND transaction_day = ?
        LIMIT 1
        """,
        (card_number, transaction_day),
    )
    return cursor.fetchone() is not None


class EmployeeCreateRequest(BaseModel):
    card_number: str = Field(min_length=1, alias="cardNumber")
    employee_id: str = Field(min_length=1, alias="employeeId")
    name: str = Field(min_length=1)
    employee_group: str = Field(min_length=1, alias="employeeGroup")
    is_disabled: Optional[bool] = Field(False, alias="isDisabled")

    class Config:
        validate_by_name = True


class EmployeeUpdateRequest(BaseModel):
    card_number: Optional[str] = Field(None, alias="cardNumber")
    name: Optional[str] = None
    employee_group: Optional[str] = Field(None, alias="employeeGroup")
    is_disabled: Optional[bool] = Field(None, alias="isDisabled")

    class Config:
        validate_by_name = True


class TenantUpdateRequest(BaseModel):
    name: str
    quota: int
    menu: List[str]
    is_limited: bool = Field(alias="isLimited")

    class Config:
        validate_by_name = True


class TenantCreateRequest(BaseModel):
    name: str
    quota: int
    menu: List[str]
    is_limited: bool = Field(alias="isLimited")

    class Config:
        validate_by_name = True


class DeviceCreateRequest(BaseModel):
    device_code: str = Field(min_length=1, alias="deviceCode")

    class Config:
        validate_by_name = True


class DeviceUpdateRequest(BaseModel):
    tenant_id: Optional[int] = Field(None, alias="tenantId")

    class Config:
        validate_by_name = True


class TransactionCreateRequest(BaseModel):
    employee_id: str = Field(min_length=1, alias="employeeId")
    tenant_id: int = Field(alias="tenantId")
    transaction_date: str = Field(min_length=1, alias="transactionDate")

    class Config:
        validate_by_name = True


class TransactionUpdateRequest(BaseModel):
    employee_id: Optional[str] = Field(None, alias="employeeId")
    tenant_id: Optional[int] = Field(None, alias="tenantId")
    transaction_date: Optional[str] = Field(None, alias="transactionDate")

    class Config:
        validate_by_name = True


TapStatusLiteral = Literal["accepted", "rejected"]
TapReasonLiteral = Literal[
    "duplicate",
    "quota_exceeded",
    "unknown_card",
    "unknown_tenant",
    "db_busy",
    "ok",
    "duplicate_daily",
    "canteen_closed",
]


class TapTransactionRequest(BaseModel):
    card_number: str = Field(min_length=1)
    tenant_id: int
    tap_ts: Optional[str] = Field(default=None, alias="tap_ts")
    tap_id: Optional[str] = Field(default=None, alias="tap_id")

    class Config:
        validate_by_name = True


class TapTransactionSummary(BaseModel):
    transaction_id: int
    card_number: str
    employee_id: str
    employee_name: str
    employee_group: Optional[str] = None
    tenant_id: int
    tenant_name: str
    transaction_date: str
    transaction_day: str


class TapTransactionResponse(BaseModel):
    status: TapStatusLiteral
    reason: TapReasonLiteral
    tap_id: Optional[str] = Field(default=None, alias="tap_id")
    server_commit_ts: Optional[int] = Field(default=None, alias="server_commit_ts")
    ticket_number: Optional[str] = None
    transaction: Optional[TapTransactionSummary] = None


class CanteenStatusUpdateRequest(BaseModel):
    mode: Literal["OPEN", "CLOSE", "NORMAL"]



def _build_tap_response(
    status_value: TapStatusLiteral,
    reason_value: TapReasonLiteral,
    *,
    ticket_number: Optional[str] = None,
    summary: Optional[TapTransactionSummary] = None,
    tap_id: Optional[str] = None,
    server_commit_ts: Optional[int] = None,
) -> TapTransactionResponse:
    return TapTransactionResponse(
        status=status_value,
        reason=reason_value,
        ticket_number=ticket_number,
        transaction=summary,
        tap_id=tap_id,
        server_commit_ts=server_commit_ts,
    )


class PreorderCreateRequest(BaseModel):
    employee_id: str = Field(min_length=1, alias="employeeId")
    tenant_id: int = Field(alias="tenantId")
    menu_label: str = Field(min_length=1, alias="menuLabel")

    class Config:
        validate_by_name = True


class PortalLoginRequest(BaseModel):
    employee_id: str = Field(min_length=1, alias="employeeId")
    portal_token: str = Field(min_length=1, alias="portalToken")

    class Config:
        validate_by_name = True


@router.post("/auth/portal-login")
def portal_login(request: PortalLoginRequest):
    """
    Entry point for authentication requests coming from the portal.
    Verifies employee_id + portal_token combination before allowing continued flow.
    """
    employee = verify_portal_token(request.employee_id, request.portal_token)
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token portal atau employee_id tidak valid.",
        )
    return {
        "employee_id": employee["employee_id"],
        "name": employee.get("name"),
        "email": employee.get("email"),
    }


@router.get("/canteen/status", status_code=status.HTTP_200_OK)
def canteen_status():
    now = datetime.datetime.now()
    status_payload = get_canteen_status(now)
    return JSONResponse(content=status_payload)


@router.get("/admin/canteen-status", status_code=status.HTTP_200_OK)
def admin_get_canteen_status():
    row = _get_or_create_canteen_status_row()
    mode = _normalize_canteen_mode(row.get("mode"))
    return {
        "mode": mode,
        "is_open": canteen_is_open(),
        "updated_at": row.get("updated_at"),
        "updated_by": row.get("updated_by"),
    }


@router.post("/admin/canteen-status", status_code=status.HTTP_200_OK)
def admin_update_canteen_status(payload: CanteenStatusUpdateRequest):
    update_canteen_mode(payload.mode, updated_by=None)
    return {"mode": payload.mode}


@router.get("/admin/check", status_code=status.HTTP_200_OK)
def admin_check(employee_id: str = Query(..., alias="employeeId")):
    trimmed_id = employee_id.strip()
    if not trimmed_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="employeeId wajib diisi.",
        )
    conn = get_db_connection()
    try:
        ensure_dashboard_admins_table(conn)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT 1 FROM dashboard_admins WHERE employee_id = ? LIMIT 1",
            (trimmed_id,),
        )
        is_admin = cursor.fetchone() is not None
        return {"employeeId": trimmed_id, "isAdmin": is_admin}
    finally:
        conn.close()


@router.get("/tenant/{tenant_id}/quota-state", status_code=status.HTTP_200_OK)
def get_tenant_quota_state(tenant_id: int):
    conn = get_db_connection()
    try:
        quota_state = evaluate_tenant_quota_for_today(conn, tenant_id)
        return JSONResponse(content=quota_state)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal menghitung kuota tenant: {exc}",
        )
    finally:
        conn.close()


@router.post("/tap", response_model=TapTransactionResponse, status_code=status.HTTP_200_OK)
def tap_transaction(tap_request: TapTransactionRequest):
    """
    Single-shot endpoint khusus device TAP.
    Melakukan lookup kartu, cek tenant serta kuota, dan menulis transaksi dalam satu transaksi DB.
    """
    start_time = time.perf_counter()
    tap_id_value = (
        (tap_request.tap_id or "").strip()
        or f"{tap_request.card_number}-{int(time.time() * 1000)}"
    )
    server_commit_ts: Optional[int] = None

    def _log_tap(status_value: TapStatusLiteral, reason_value: TapReasonLiteral):
        duration_ms = (time.perf_counter() - start_time) * 1000
        print(
            f"[tap_transaction] tap_id={tap_id_value} status={status_value} reason={reason_value} duration={duration_ms:.2f}ms"
        )

    def _trace_stage(stage: str) -> int:
        timestamp_ms = int(time.time() * 1000)
        print(f"[tap_trace] tap_id={tap_id_value} stage={stage} ts={timestamp_ms}")
        return timestamp_ms

    def _tap_response(
        status_value: TapStatusLiteral,
        reason_value: TapReasonLiteral,
        *,
        summary: Optional[TapTransactionSummary] = None,
    ) -> TapTransactionResponse:
        response = _build_tap_response(
            status_value=status_value,
            reason_value=reason_value,
            summary=summary,
            tap_id=tap_id_value,
            server_commit_ts=server_commit_ts,
        )
        _log_tap(status_value, reason_value)
        return response

    _trace_stage("t_server_start")

    _, transaction_date_text, transaction_day = _normalize_timestamp_to_local(
        tap_request.tap_ts, field_name="tap_ts"
    )
    conn = get_db_connection(mode="tap")
    cursor = conn.cursor()
    summary: Optional[TapTransactionSummary] = None
    employee_row = None
    tenant_row = None

    def _rollback_if_needed():
        if conn.in_transaction:
            conn.rollback()

    try:
        cursor.execute(
            """
            SELECT employee_id,
                   card_number,
                   name,
                   employee_group,
                   is_disabled,
                   is_blocked
            FROM employees
            WHERE card_number = ?
            """,
            (tap_request.card_number,),
        )
        employee = cursor.fetchone()
        if (
            not employee
            or employee["is_disabled"]
            or employee["is_blocked"]
        ):
            return _tap_response(
                status_value="rejected", reason_value="unknown_card"
            )
        employee_row = employee

        cursor.execute(
            """
            SELECT id, name, quota, is_limited
            FROM tenants
            WHERE id = ?
            """,
            (tap_request.tenant_id,),
        )
        tenant = cursor.fetchone()
        if not tenant:
            return _tap_response(
                status_value="rejected", reason_value="unknown_tenant"
            )
        tenant_row = tenant
        quota_value = tenant["quota"] or 0

        _trace_stage("t_before_begin_immediate")
        cursor.execute("BEGIN IMMEDIATE")
        _trace_stage("t_after_begin_immediate")

        if _card_has_transaction_for_day(
            cursor, employee["card_number"], transaction_day
        ):
            _rollback_if_needed()
            return _tap_response(
                status_value="rejected", reason_value="duplicate_daily"
            )

        cursor.execute(
            """
            SELECT COUNT(*)
            FROM transactions
            WHERE tenant_id = ?
              AND transaction_day = ?
            """,
            (tenant["id"], transaction_day),
        )
        tenant_orders_today = cursor.fetchone()[0]
        remaining_slots = quota_value - tenant_orders_today
        allow_due_to_free_mode = False
        if quota_value == 0:
            quota_reason = "unlimited"
        elif remaining_slots > 0:
            quota_reason = "ok"
        else:
            cursor.execute(
                """
                SELECT t.id,
                       t.quota,
                       COALESCE(tx.order_count, 0) AS order_count
                FROM tenants t
                LEFT JOIN (
                    SELECT tenant_id, COUNT(*) AS order_count
                    FROM transactions
                    WHERE transaction_day = ?
                    GROUP BY tenant_id
                ) tx ON tx.tenant_id = t.id
                WHERE t.quota IS NOT NULL AND t.quota > 0
                """,
                (transaction_day,),
            )
            tenants_with_quota = cursor.fetchall()
            any_tenant_with_remaining = False
            for tenant_info in tenants_with_quota:
                tenant_quota = tenant_info["quota"] or 0
                tenant_taken = tenant_info["order_count"] or 0
                if tenant_quota - tenant_taken > 0:
                    any_tenant_with_remaining = True
                    break
            if any_tenant_with_remaining:
                quota_reason = "quota_exceeded"
            else:
                quota_reason = "free_mode"
                allow_due_to_free_mode = True
        print(
            f"[tap_quota] tenant_id={tenant['id']} quota={quota_value} count_today={tenant_orders_today} remaining={remaining_slots} reason={quota_reason}"
        )
        if quota_value > 0 and remaining_slots <= 0 and not allow_due_to_free_mode:
            _rollback_if_needed()
            return _tap_response(
                status_value="rejected", reason_value="quota_exceeded"
            )

        _trace_stage("t_before_insert")
        cursor.execute(
            """
            INSERT INTO transactions (
                card_number,
                employee_id,
                employee_name,
                employee_group,
                tenant_id,
                tenant_name,
                transaction_date,
                transaction_day
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                employee["card_number"],
                employee["employee_id"],
                employee["name"],
                employee["employee_group"],
                tenant["id"],
                tenant["name"],
                transaction_date_text,
                transaction_day,
            ),
        )
        transaction_id = cursor.lastrowid
        conn.commit()
        server_commit_ts = _trace_stage("t_after_commit")

        summary = TapTransactionSummary(
            transaction_id=transaction_id,
            card_number=employee_row["card_number"],
            employee_id=employee_row["employee_id"],
            employee_name=employee_row["name"],
            employee_group=employee_row["employee_group"],
            tenant_id=tenant_row["id"],
            tenant_name=tenant_row["name"],
            transaction_date=transaction_date_text,
            transaction_day=transaction_day,
        )
    except sqlite3.IntegrityError as exc:
        _rollback_if_needed()
        error_text = str(exc).lower()
        if "unique" in error_text or "transaction_day" in error_text:
            return _tap_response(
                status_value="rejected", reason_value="duplicate_daily"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal membuat transaksi TAP: {exc}",
        )
    except sqlite3.OperationalError as exc:
        _rollback_if_needed()
        message = str(exc).lower()
        if "database is locked" in message:
            resp = _tap_response(status_value="rejected", reason_value="db_busy")
            return JSONResponse(
                status_code=status.HTTP_409_CONFLICT,
                content=resp.dict(),
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error saat TAP: {exc}",
        )
    except sqlite3.IntegrityError as exc:
        _rollback_if_needed()
        error_text = str(exc).lower()
        if (
            "transaction_day" in error_text
            or "idx_transactions_card_day" in error_text
            or "unique" in error_text
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=DAILY_TRANSACTION_LIMIT_MESSAGE,
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal membuat pre-order: {exc}",
        )
    except HTTPException:
        _rollback_if_needed()
        raise
    except Exception as exc:
        _rollback_if_needed()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"TAP endpoint gagal: {exc}",
        )
    finally:
        conn.close()

    if not summary or not employee_row or not tenant_row:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="TAP berhasil tetapi ringkasan tidak terbentuk.",
        )

    response_payload = _tap_response(
        status_value="accepted",
        reason_value="ok",
        summary=summary,
    )

    _trace_stage("t_before_sse_trigger")
    if sse_manager.main_event_loop:
        sse_payload = {
            "id": str(tenant_row["id"]),
            "name": employee_row["name"],
            "employee_id": employee_row["employee_id"],
            "tap_id": tap_id_value,
            "server_commit_ts": server_commit_ts,
        }
        asyncio.run_coroutine_threadsafe(
            sse_manager.trigger_sse_event_async(sse_payload),
            sse_manager.main_event_loop,
        )
    _trace_stage("t_after_sse_trigger")

    return response_payload


@router.post("/transaction", status_code=status.HTTP_201_CREATED)
def create_transaction(transaction_data: TransactionCreateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    start_time = time.perf_counter()
    success = False
    print(
        f"[create_transaction] start tenant_id={transaction_data.tenant_id} employee_id={transaction_data.employee_id}"
    )
    try:

        cursor.execute(
            "SELECT card_number, name, employee_group FROM employees WHERE employee_id = ?",
            (transaction_data.employee_id,),
        )
        employee = cursor.fetchone()
        if not employee:
            raise HTTPException(
                status_code=404,
                detail=f"Employee with ID '{transaction_data.employee_id}' not found",
            )

        cursor.execute(
            "SELECT name FROM tenants WHERE id = ?", (transaction_data.tenant_id,)
        )
        tenant = cursor.fetchone()
        if not tenant:
            raise HTTPException(
                status_code=404,
                detail=f"Tenant with ID '{transaction_data.tenant_id}' not found",
            )

        employee_card_number = employee["card_number"]
        employee_name = employee["name"]
        employee_group = employee["employee_group"]
        tenant_name = tenant["name"]

        (
            _,
            normalized_transaction_date,
            transaction_day,
        ) = _normalize_timestamp_to_local(
            transaction_data.transaction_date, field_name="transactionDate"
        )

        if _card_has_transaction_for_day(cursor, employee_card_number, transaction_day):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=DAILY_TRANSACTION_LIMIT_MESSAGE,
            )

        cursor.execute(
            """
            INSERT INTO transactions (
                card_number,
                employee_id,
                employee_name,
                employee_group,
                tenant_id,
                tenant_name,
                transaction_date,
                transaction_day
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                employee_card_number,
                transaction_data.employee_id,
                employee_name,
                employee_group,
                transaction_data.tenant_id,
                tenant_name,
                normalized_transaction_date,
                transaction_day,
            ),
        )
        conn.commit()
        server_commit_ts = int(time.time() * 1000)

        sse_payload = {
            "id": str(transaction_data.tenant_id),
            "name": employee_name,
            "employee_id": transaction_data.employee_id,
            "tap_id": None,
            "server_commit_ts": server_commit_ts,
        }
        if sse_manager.main_event_loop:
            asyncio.run_coroutine_threadsafe(
                sse_manager.trigger_sse_event_async(sse_payload),
                sse_manager.main_event_loop,
            )

        success = True
        return {"message": "Transaction logged successfully."}
    except sqlite3.IntegrityError as exc:
        conn.rollback()
        error_text = str(exc).lower()
        if "unique" in error_text or "transaction_day" in error_text:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=DAILY_TRANSACTION_LIMIT_MESSAGE,
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to log transaction: {exc}",
        )
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to log transaction: {e}",
        )
    finally:
        conn.close()
        duration_ms = (time.perf_counter() - start_time) * 1000
        status_label = "ok" if success else "failed"
        print(
            f"[create_transaction] end tenant_id={transaction_data.tenant_id} employee_id={transaction_data.employee_id} status={status_label} duration={duration_ms:.2f}ms"
        )


@router.post("/preorder", status_code=status.HTTP_201_CREATED)
def create_preorder(preorder: PreorderCreateRequest, background_tasks: BackgroundTasks):
    now_dt = datetime.datetime.now()
    if not canteen_is_open(now_dt):
        canteen_status = get_canteen_status(now_dt)
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=canteen_status["message"],
        )

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        conn.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute(
            """
            SELECT employee_id, card_number, name, is_disabled, is_blocked, email, employee_group
            FROM employees
            WHERE employee_id = ?
            """,
            (preorder.employee_id,),
        )
        employee = cursor.fetchone()
        if (
            not employee
            or employee["is_disabled"]
            or employee["is_blocked"]
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Pegawai tidak dapat melakukan pre-order.",
                )

        card_number = employee["card_number"] or ""
        today = datetime.datetime.now(tz=JAKARTA_TZ).date().isoformat()

        if _card_has_transaction_for_day(cursor, card_number, today):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=DAILY_TRANSACTION_LIMIT_MESSAGE,
            )

        employee_email = (employee["email"] or "").strip()
        if not employee_email:
            employee_email = update_employee_email(preorder.employee_id)
            if not employee_email:
                print(
                    f"Peringatan: email untuk employee_id {preorder.employee_id} tidak ditemukan di DB maupun dummy mapping."
                )

        cursor.execute(
            "SELECT id, name, quota, is_limited, verification_code FROM tenants WHERE id = ?",
            (preorder.tenant_id,),
        )
        tenant = cursor.fetchone()
        if not tenant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Tenant dengan ID '{preorder.tenant_id}' tidak ditemukan.",
            )

        tenant_prefix = get_tenant_prefix(tenant["name"])
        try:
            quota_state = evaluate_tenant_quota_for_today(conn, tenant["id"])
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            )

        if not quota_state["can_order_for_target"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Kuota tenant ini sudah habis sementara tenant lain masih memiliki sisa.",
            )

        day_start, day_end = _get_local_day_bounds_from_string(today)
        cursor.execute(
            """
            SELECT 1
            FROM preorders
            WHERE employee_id = ?
              AND order_date = ?
            LIMIT 1
            """,
            (preorder.employee_id, today),
        )
        if cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Anda sudah melakukan pre-order hari ini. Hanya satu order per hari yang diperbolehkan.",
            )

        cursor.execute(
            """
            SELECT id
            FROM preorders
            WHERE employee_id = ?
              AND tenant_id = ?
              AND order_date = ?
              AND menu_label = ?
              AND created_at >= datetime('now','localtime','-15 seconds')
            ORDER BY id DESC
            LIMIT 1
            """,
            (
                preorder.employee_id,
                preorder.tenant_id,
                today,
                preorder.menu_label,
            ),
        )
        if cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Pesanan ini sudah tercatat. Kemungkinan Anda menekan tombol dua kali atau membuka lebih dari satu tab.",
            )

        tenant_verification_code = tenant["verification_code"]

        cursor.execute(
            """
            SELECT COUNT(*)
            FROM transactions
            WHERE tenant_id = ?
              AND transaction_date >= ?
              AND transaction_date < ?
            """,
            (preorder.tenant_id, day_start, day_end),
        )
        order_count_today = cursor.fetchone()[0] + 1
        transaction_number = order_count_today

        quota_value = tenant["quota"] or 0

        if quota_value > 0:
            remaining_before_new_order = quota_value - (order_count_today - 1)
            if remaining_before_new_order <= 0:
                cursor.execute(
                    """
                    SELECT t.id,
                           t.quota,
                           COALESCE(tx.order_count, 0) AS order_count
                    FROM tenants t
                    LEFT JOIN (
                        SELECT tenant_id, COUNT(*) AS order_count
                        FROM transactions
                        WHERE transaction_date >= ?
                          AND transaction_date < ?
                        GROUP BY tenant_id
                    ) tx ON tx.tenant_id = t.id
                    WHERE t.quota IS NOT NULL AND t.quota > 0
                    """,
                    (day_start, day_end),
                )
                tenants_with_quota = cursor.fetchall()
                any_tenant_with_remaining = False
                for tenant_row in tenants_with_quota:
                    tenant_quota = tenant_row["quota"] or 0
                    orders_taken = tenant_row["order_count"] or 0
                    if tenant_quota - orders_taken > 0:
                        any_tenant_with_remaining = True
                        break

                if any_tenant_with_remaining:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Kuota tenant ini sudah habis. Silakan pilih tenant lain yang masih tersedia.",
                    )
        remaining_after = quota_value - order_count_today

        if quota_value > 0:
            if remaining_after >= 0:
                queue_number = remaining_after + 1
            else:
                queue_number = remaining_after
        else:
            queue_number = remaining_after

        if queue_number is None:
            queue_code = "-"
        elif tenant_prefix:
            queue_code = f"{tenant_prefix}{queue_number}"
        else:
            queue_code = str(queue_number)

        remaining_quota: Optional[int] = None
        if quota_value > 0:
            remaining_quota = remaining_after

        order_code = uuid4().hex
        order_datetime = datetime.datetime.now(tz=JAKARTA_TZ)
        transaction_timestamp = order_datetime.strftime("%Y-%m-%d %H:%M:%S")
        transaction_day = today
        ticket_number = generate_ticket_number(order_datetime, transaction_number)
        weekday_names = [
            "Senin",
            "Selasa",
            "Rabu",
            "Kamis",
            "Jumat",
            "Sabtu",
            "Minggu",
        ]
        weekday_name = weekday_names[order_datetime.weekday()]
        order_datetime_text = f"{weekday_name} {order_datetime.strftime('%d/%m/%Y, %H.%M')}"
        cursor.execute(
            """
            INSERT INTO preorders (
                order_code,
                ticket_number,
                employee_id,
                card_number,
                employee_name,
                tenant_id,
                tenant_name,
                menu_label,
                order_date,
                status,
                queue_number
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'PENDING', ?)
            """,
            (
                order_code,
                ticket_number,
                preorder.employee_id,
                card_number,
                employee["name"],
                tenant["id"],
                tenant["name"],
                preorder.menu_label,
                today,
                queue_number,
            ),
        )
        cursor.execute(
            """
            INSERT INTO transactions (
                card_number,
                employee_id,
                employee_name,
                employee_group,
                tenant_id,
                tenant_name,
                transaction_date,
                transaction_day
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                card_number,
                preorder.employee_id,
                employee["name"],
                employee["employee_group"],
                tenant["id"],
                tenant["name"],
                transaction_timestamp,
                transaction_day,
            ),
        )
        conn.commit()
        server_commit_ts = int(time.time() * 1000)

        # Trigger SSE to update monitoring dashboard after pre-order/transaction creation
        if sse_manager.main_event_loop:
            sse_payload = {
                "id": str(tenant["id"]),
                "name": employee["name"],
                "employee_id": preorder.employee_id,
                "tap_id": None,
                "server_commit_ts": server_commit_ts,
            }
            asyncio.run_coroutine_threadsafe(
                sse_manager.trigger_sse_event_async(sse_payload),
                sse_manager.main_event_loop,
            )

        order_payload = {
            "ticket_number": ticket_number,
            "order_code": order_code,
            "employee_id": preorder.employee_id,
            "employee_name": employee["name"],
            "tenant_name": tenant["name"],
            "tenant_verification_code": tenant_verification_code,
            "order_datetime_text": order_datetime_text,
            "order_date": today,
            "menu_label": preorder.menu_label,
            "menu_items": [{"label": preorder.menu_label, "qty": 1}],
            "queue_number": queue_number,
            "queue_code": queue_code,
            "transaction_number": ticket_number,
        }
        if employee_email:
            background_tasks.add_task(
                send_order_confirmation,
                employee_email,
                order_payload,
                recipient="employee",
            )
        canteen_email = os.getenv("CANTEEN_ORDER_EMAIL")
        if canteen_email:
            background_tasks.add_task(
                send_order_confirmation,
                canteen_email,
                order_payload,
                recipient="canteen",
            )

        return JSONResponse(
            content={
                "orderCode": ticket_number,
                "orderHash": order_code,
                "employeeId": preorder.employee_id,
                "employeeName": employee["name"],
                "tenantId": tenant["id"],
                "tenantName": tenant["name"],
                "tenantVerificationCode": tenant_verification_code,
                "menuLabel": preorder.menu_label,
                "orderDate": today,
                "orderDateTimeText": order_datetime_text,
                "queueNumber": queue_number,
                "queueCode": queue_code,
                "transactionNumber": ticket_number,
                "remainingQuota": remaining_quota,
            }
        )
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal membuat pre-order: {e}",
        )
    finally:
        conn.close()


# DEPRECATED: Endpoint admin sekali-jalan untuk mengirim token karyawan via email.
# Flow generate_tokens / broadcast_employee_tokens tidak lagi dipakai oleh autentikasi utama.
@router.post("/admin/send-employee-tokens", status_code=status.HTTP_410_GONE)
def send_employee_tokens():
    raise HTTPException(
        status_code=status.HTTP_410_GONE,
        detail="DEPRECATED: Token email tidak lagi digunakan dalam alur Cawang Canteen.",
    )


@router.get("/transaction/check_duplicate", status_code=status.HTTP_200_OK)
async def check_duplicate_transaction(
    card_number: str = Query(..., alias="cardNumber"),
    transaction_date: str = Query(..., alias="transactionDate"),
):
    conn = get_db_connection()
    cursor = conn.cursor()
    day_start, day_end = _get_local_day_bounds_from_string(transaction_date)
    try:
        cursor.execute(
            """
            SELECT 1 FROM transactions
            WHERE card_number = ?
              AND transaction_date >= ?
              AND transaction_date < ?
            """,
            (card_number, day_start, day_end),
        )
        exists = cursor.fetchone() is not None
        return {"exists": exists}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to check duplicate transaction: {e}",
        )
    finally:
        conn.close()


@router.get("/transaction/daily_count", status_code=status.HTTP_200_OK)
async def get_daily_transaction_count(
    tenant_id: int = Query(..., alias="tenantId"),
    transaction_date: str = Query(..., alias="transactionDate"),
):
    conn = get_db_connection()
    cursor = conn.cursor()
    day_start, day_end = _get_local_day_bounds_from_string(transaction_date)
    try:
        cursor.execute(
            """
            SELECT COUNT(*)
            FROM transactions
            WHERE tenant_id = ?
              AND transaction_date >= ?
              AND transaction_date < ?
            """,
            (tenant_id, day_start, day_end),
        )
        count = cursor.fetchone()[0]
        return {"count": count}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get daily transaction count: {e}",
        )
    finally:
        conn.close()


@router.get("/transaction/{transaction_id}/detail", status_code=status.HTTP_200_OK)
def get_transaction(transaction_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM transactions WHERE id = ?", (transaction_id,))
        transaction = cursor.fetchone()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        return JSONResponse(content=dict(transaction))
    finally:
        conn.close()


@router.put("/transaction/{transaction_id}/update", status_code=status.HTTP_200_OK)
def update_transaction(transaction_id: int, transaction_data: TransactionUpdateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        update_data_dict = transaction_data.dict(by_alias=True, exclude_unset=True)
        if not update_data_dict:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="No fields to update"
            )

        update_fields = {}

        if "employeeId" in update_data_dict:
            employee_id = update_data_dict["employeeId"]
            cursor.execute(
                "SELECT card_number, name, employee_group FROM employees WHERE employee_id = ?",
                (employee_id,),
            )
            employee = cursor.fetchone()
            if not employee:
                raise HTTPException(
                    status_code=404,
                    detail=f"Employee with ID '{employee_id}' not found",
                )
            update_fields["employee_id"] = employee_id
            update_fields["card_number"] = employee["card_number"]
            update_fields["employee_name"] = employee["name"]
            update_fields["employee_group"] = employee["employee_group"]

        if "tenantId" in update_data_dict:
            tenant_id = update_data_dict["tenantId"]
            cursor.execute("SELECT name FROM tenants WHERE id = ?", (tenant_id,))
            tenant = cursor.fetchone()
            if not tenant:
                raise HTTPException(
                    status_code=404,
                    detail=f"Tenant with ID '{tenant_id}' not found",
                )
            update_fields["tenant_id"] = tenant_id
            update_fields["tenant_name"] = tenant["name"]

        if "transactionDate" in update_data_dict:
            update_fields["transaction_date"] = update_data_dict["transactionDate"]

        if not update_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid fields to update",
            )

        set_clause = ", ".join([f"{key} = ?" for key in update_fields.keys()])
        query = f"UPDATE transactions SET {set_clause} WHERE id = ?"
        params = list(update_fields.values()) + [transaction_id]

        cursor.execute(query, tuple(params))
        conn.commit()

        return {"message": "Transaction updated successfully."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update transaction: {e}",
        )
    finally:
        conn.close()


@router.delete("/transaction/{transaction_id}/delete", status_code=status.HTTP_200_OK)
def delete_transaction(transaction_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM transactions WHERE id = ?", (transaction_id,))
        conn.commit()
        return {"message": "Transaction deleted successfully."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete transaction: {e}",
        )
    finally:
        conn.close()


@router.delete("/device/all", status_code=status.HTTP_200_OK)
async def delete_all_devices():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM devices")
        conn.commit()
        return {"message": "All devices deleted successfully."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete all devices: {e}",
        )
    finally:
        conn.close()


@router.post("/device/setup", status_code=status.HTTP_200_OK)
async def setup_devices_table():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_code TEXT NOT NULL UNIQUE,
                tenant_id INTEGER,
                FOREIGN KEY (tenant_id) REFERENCES tenants (id)
            );
        """
        )
        conn.commit()
        return {"message": "Devices table ensured to exist."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to setup devices table: {e}",
        )
    finally:
        conn.close()


@router.get("/device/check/{device_code}", status_code=status.HTTP_200_OK)
async def check_device_exists(device_code: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM devices WHERE device_code = ?", (device_code,))
        exists = cursor.fetchone() is not None
        return {"exists": exists}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to check device existence: {e}",
        )
    finally:
        conn.close()


@router.post("/device", status_code=status.HTTP_201_CREATED)
async def register_new_device(device_data: DeviceCreateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT 1 FROM devices WHERE device_code = ?", (device_data.device_code,)
        )
        if cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Device '{device_data.device_code}' already registered.",
            )

        cursor.execute(
            "INSERT INTO devices (device_code) VALUES (?)", (device_data.device_code,)
        )
        conn.commit()
        return {
            "message": f"Device '{device_data.device_code}' registered successfully."
        }
    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to register device: {e}",
        )
    finally:
        conn.close()


@router.get("/sse")
async def sse_endpoint(request: Request):
    return StreamingResponse(
        sse_manager.event_stream(request),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        },
    )


@router.get("/employee")
async def get_employees():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM employees ORDER BY name")
        employees = [dict(row) for row in cursor.fetchall()]
        return JSONResponse(content=employees)
    finally:
        conn.close()


@router.get("/tenant")
async def get_tenants():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM tenants ORDER BY name")
        tenants_list = [dict(row) for row in cursor.fetchall()]

        for tenant in tenants_list:
            cursor.execute(
                "SELECT menu FROM tenant_menu WHERE tenant_id = ?", (tenant["id"],)
            )
            menus = cursor.fetchall()
            tenant["menu"] = [row["menu"] for row in menus]

        return JSONResponse(content=tenants_list)
    finally:
        conn.close()


@router.get("/device")
async def get_devices():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:

        cursor.execute(
            """
            SELECT
                d.device_code,
                t.id AS tenant_id,
                t.name AS tenant_name
            FROM devices d
            LEFT JOIN tenants t ON d.tenant_id = t.id
            ORDER BY d.device_code
        """
        )

        devices = []
        for row in cursor.fetchall():
            tenant_info = None
            if row["tenant_id"]:
                tenant_info = {"id": row["tenant_id"], "name": row["tenant_name"]}
            devices.append({"device_code": row["device_code"], "tenant": tenant_info})

        return JSONResponse(content=devices)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch devices: {e}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/dashboard/overview", status_code=status.HTTP_200_OK)
def get_dashboard_overview():
    """
    Provides a dashboard overview of all devices with assigned tenants,
    including tenant details, menu, and today's transaction count.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        today = datetime.date.today().isoformat()
        day_start, day_end = _get_local_day_bounds_from_string()

        query = """
            SELECT
                d.device_code,
                t.id AS tenant_id,
                t.name AS tenant_name,
                COALESCE(t.quota, 0) AS quota,
                t.verification_code
            FROM
                devices d
            JOIN
                tenants t ON d.tenant_id = t.id
            WHERE
                d.tenant_id IS NOT NULL
            ORDER BY
                d.tenant_id
        """
        cursor.execute(query)

        devices_with_tenants = cursor.fetchall()

        result = []
        tenant_ids = list(set(row["tenant_id"] for row in devices_with_tenants))

        all_menus = {}
        if tenant_ids:
            placeholders = ",".join("?" for _ in tenant_ids)
            menu_query = f"SELECT tenant_id, menu FROM tenant_menu WHERE tenant_id IN ({placeholders})"
            cursor.execute(menu_query, tenant_ids)

            for menu_row in cursor.fetchall():
                tid = menu_row["tenant_id"]
                if tid not in all_menus:
                    all_menus[tid] = []
                all_menus[tid].append(menu_row["menu"])

        for row in devices_with_tenants:
            tenant_id = row["tenant_id"]

            cursor.execute(
                """
                SELECT COALESCE(COUNT(*), 0) AS ordered_count
                FROM transactions
                WHERE tenant_id = ?
                  AND transaction_date >= ?
                  AND transaction_date < ?
                """,
                (tenant_id, day_start, day_end),
            )
            ordered_row = cursor.fetchone()
            ordered = int(ordered_row["ordered_count"]) if ordered_row else 0

            cursor.execute(
                """
                SELECT
                    tr.employee_name,
                    tr.employee_id,
                    tr.transaction_date,
                    tr.id
                FROM transactions tr
                WHERE tr.tenant_id = ?
                  AND tr.transaction_date >= ?
                  AND tr.transaction_date < ?
                ORDER BY tr.transaction_date DESC, tr.id DESC
                LIMIT 1
                """,
                (tenant_id, day_start, day_end),
            )
            last_order_row = cursor.fetchone()
            last_order = None
            if last_order_row:
                last_order = {
                    "queueNumber": None,
                    "menuLabel": None,
                    "employeeName": last_order_row["employee_name"],
                    "employeeId": last_order_row["employee_id"],
                }

            quota_value = int(row["quota"] or 0)
            available = quota_value - ordered

            verification_code = row["verification_code"]

            device_info = {
                "device_code": row["device_code"],
                "tenantId": tenant_id,
                "tenantName": row["tenant_name"],
                "available": available,
                "ordered": ordered,
                "lastOrder": last_order,
                "tenantVerificationCode": verification_code,
                "tenant": {
                    "id": tenant_id,
                    "name": row["tenant_name"],
                    "menu": all_menus.get(tenant_id, []),
                    "quota": quota_value,
                    "ordered": ordered,
                    "available": available,
                    "lastOrder": last_order,
                    "verificationCode": verification_code,
                },
            }
            result.append(device_info)

        return JSONResponse(content=result)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch dashboard overview: {e}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/device/assigned")
async def get_assigned_devices_with_menu():
    """
    Returns a list of all devices that have an assigned tenant,
    including the tenant's full details and menu.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT
                d.device_code,
                t.id AS tenant_id,
                t.name AS tenant_name,
                t.quota,
                t.is_limited,
                tm.menu
            FROM
                devices d
            INNER JOIN
                tenants t ON d.tenant_id = t.id
            LEFT JOIN
                tenant_menu tm ON t.id = tm.tenant_id
            WHERE
                d.tenant_id IS NOT NULL
            ORDER BY
                d.device_code, t.id
        """
        )

        assigned_devices = {}
        for row in cursor.fetchall():
            device_code = row["device_code"]

            if device_code not in assigned_devices:
                assigned_devices[device_code] = {
                    "device_code": device_code,
                    "tenant": {
                        "id": row["tenant_id"],
                        "name": row["tenant_name"],
                        "quota": row["quota"],
                        "is_limited": row["is_limited"],
                        "menu": [],
                    },
                }

            if row["menu"] is not None:
                assigned_devices[device_code]["tenant"]["menu"].append(row["menu"])

        return JSONResponse(content=list(assigned_devices.values()))
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch assigned devices: {e}",
        )
    finally:
        if conn:
            conn.close()


@router.put("/device/{device_code}")
async def update_device_tenant(device_code: str, update_data: DeviceUpdateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE devices SET tenant_id = ? WHERE device_code = ?",
            (update_data.tenant_id, device_code),
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update device: {e}",
        )
    finally:
        conn.close()

    return JSONResponse(
        content={"message": f"Device '{device_code}' updated successfully."},
        status_code=status.HTTP_200_OK,
    )


@router.post("/tenant", status_code=status.HTTP_201_CREATED)
async def create_tenant(create_data: TenantCreateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        verification_code = generate_verification_code()
        cursor.execute(
            "INSERT INTO tenants (name, quota, is_limited, verification_code) VALUES (?, ?, ?, ?)",
            (
                create_data.name,
                create_data.quota,
                create_data.is_limited,
                verification_code,
            ),
        )
        new_tenant_id = cursor.lastrowid

        if create_data.menu:
            menu_items = [(new_tenant_id, item) for item in create_data.menu]
            cursor.executemany(
                "INSERT INTO tenant_menu (tenant_id, menu) VALUES (?, ?)", menu_items
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
    return {
        "message": "Tenant created successfully",
        "tenantId": new_tenant_id,
        "verificationCode": verification_code,
    }


@router.put("/tenant/{tenant_id}/update")
async def update_tenant(tenant_id: int, update_data: TenantUpdateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT verification_code FROM tenants WHERE id = ?",
            (tenant_id,),
        )
        row = cursor.fetchone()
        verification_code = row["verification_code"] if row else None

        cursor.execute(
            "UPDATE tenants SET name = ?, quota = ?, is_limited = ? WHERE id = ?",
            (
                update_data.name,
                update_data.quota,
                update_data.is_limited,
                tenant_id,
            ),
        )

        cursor.execute("DELETE FROM tenant_menu WHERE tenant_id = ?", (tenant_id,))

        if update_data.menu:
            menu_items = [(tenant_id, item) for item in update_data.menu]
            cursor.executemany(
                "INSERT INTO tenant_menu (tenant_id, menu) VALUES (?, ?)", menu_items
            )

        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
    return {
        "message": "Tenant updated successfully",
        "verificationCode": verification_code,
    }


@router.get("/tenant/{tenant_id}/detail")
async def get_tenant(tenant_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM tenants WHERE id = ?", (tenant_id,))
        tenant = cursor.fetchone()
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant not found")

        new_code = generate_verification_code()
        cursor.execute(
            "UPDATE tenants SET verification_code = ? WHERE id = ?",
            (new_code, tenant_id),
        )

        tenant_dict = dict(tenant)
        tenant_dict["verification_code"] = new_code

        cursor.execute("SELECT menu FROM tenant_menu WHERE tenant_id = ?", (tenant_id,))
        menus = cursor.fetchall()
        tenant_dict["menu"] = [row["menu"] for row in menus]

        conn.commit()
        return tenant_dict
    finally:
        conn.close()


@router.delete("/tenant/{tenant_id}/delete", status_code=status.HTTP_200_OK)
async def delete_tenant(tenant_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE devices SET tenant_id = NULL WHERE tenant_id = ?", (tenant_id,)
        )
        cursor.execute("DELETE FROM tenant_menu WHERE tenant_id = ?", (tenant_id,))
        cursor.execute("DELETE FROM tenants WHERE id = ?", (tenant_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
    return {"message": f"Tenant with id '{tenant_id}' deleted successfully."}


@router.post("/employee", status_code=status.HTTP_201_CREATED)
async def create_employee(create_data: EmployeeCreateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT * FROM employees WHERE employee_id = ? OR card_number = ?",
            (create_data.employee_id, create_data.card_number),
        )
        if cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Employee with ID '{create_data.employee_id}' or card number '{create_data.card_number}' already exists.",
            )

        cursor.execute(
            """
            INSERT INTO employees (employee_id, card_number, name, employee_group, admin, is_disabled)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                create_data.employee_id,
                create_data.card_number,
                create_data.name,
                create_data.employee_group,
                False,
                create_data.is_disabled,
            ),
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create employee: {e}",
        )
    finally:
        conn.close()

    return JSONResponse(
        content={
            "message": f"Employee with ID '{create_data.employee_id}' created successfully."
        },
        status_code=status.HTTP_201_CREATED,
    )


@router.get("/employee/{employee_id}/detail")
async def get_employee(employee_id: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM employees WHERE employee_id = ?", (employee_id,))
    employee = cursor.fetchone()
    conn.close()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Employee with ID '{employee_id}' not found.",
        )
    return JSONResponse(content=dict(employee))


@router.put("/employee/{employee_id}/update")
async def update_employee(employee_id: str, update_data: EmployeeUpdateRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM employees WHERE employee_id = ?", (employee_id,))
        if not cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee with ID '{employee_id}' not found.",
            )

        update_fields = {k: v for k, v in update_data.dict(exclude_unset=True).items()}
        if not update_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="No fields to update"
            )

        set_clause = ", ".join([f"{key} = ?" for key in update_fields.keys()])
        query = f"UPDATE employees SET {set_clause} WHERE employee_id = ?"
        params = list(update_fields.values()) + [employee_id]

        cursor.execute(query, tuple(params))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update employee: {e}",
        )
    finally:
        conn.close()

    return JSONResponse(
        content={"message": f"Employee with ID '{employee_id}' updated successfully."},
        status_code=status.HTTP_200_OK,
    )


@router.delete("/employee/{employee_id}/delete", status_code=status.HTTP_200_OK)
async def delete_employee(employee_id: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM employees WHERE employee_id = ?", (employee_id,))
        if not cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee with ID '{employee_id}' not found.",
            )

        cursor.execute("DELETE FROM employees WHERE employee_id = ?", (employee_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete employee: {e}",
        )
    finally:
        conn.close()

    return {"message": f"Employee with ID '{employee_id}' deleted successfully."}


@router.get("/transaction/report")
async def get_transaction_reports(
    search: Optional[str] = Query(
        None,
        description="Search by employee name, tenant name, or card number (case-insensitive, partial match)",
    ),
    employee_group: Optional[str] = Query(
        None,
        description="Filter by employee group (case-insensitive, partial match)",
    ),
    employee: Optional[str] = Query(
        None,
        description="Filter by employee name or employee ID (case-insensitive, partial match)",
    ),
    tenant: Optional[str] = Query(
        None,
        description="Filter by tenant name (case-insensitive, partial match)",
    ),
    start_date: Optional[str] = Query(
        None, description="Start date for period filter (YYYY-MM-DD)"
    ),
    end_date: Optional[str] = Query(
        None, description="End date for period filter (YYYY-MM-DD)"
    ),
    page: int = Query(1, ge=1, description="Page number for pagination"),
    page_size: int = Query(10, ge=1, le=100, description="Number of items per page"),
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        query_params = []
        where_clauses = []

        if search:
            where_clauses.append(
                "(LOWER(employee_name) LIKE ? OR LOWER(tenant_name) LIKE ? OR card_number LIKE ?)"
            )
            query_params.extend(
                [f"%{search.lower()}%", f"%{search.lower()}%", f"%{search}%"]
            )
        if employee_group:
            where_clauses.append("(LOWER(employee_group) LIKE ?)")
            query_params.append(f"%{employee_group.lower()}%")
        if employee:
            where_clauses.append("(employee_id LIKE ? OR LOWER(employee_name) LIKE ?)")
            query_params.append(f"%{employee.lower()}%")
            query_params.append(f"%{employee.lower()}%")
        if tenant:
            where_clauses.append("(LOWER(tenant_name) LIKE ?)")
            query_params.append(f"%{tenant.lower()}%")
        if start_date:
            where_clauses.append("transaction_date >= ?")
            query_params.append(start_date)
        if end_date:
            where_clauses.append("transaction_date <= ?")
            query_params.append(end_date)

        where_sql = ""
        if where_clauses:
            where_sql = "WHERE " + " AND ".join(where_clauses)

        count_query = f"SELECT COUNT(*) FROM transactions {where_sql}"
        cursor.execute(count_query, tuple(query_params))
        total_items = cursor.fetchone()[0]
        total_pages = math.ceil(total_items / page_size) if total_items > 0 else 0

        offset = (page - 1) * page_size
        data_query = f"""
            SELECT id, card_number, employee_id, employee_name, employee_group, tenant_id, tenant_name, transaction_date
            FROM transactions
            {where_sql}
            ORDER BY transaction_date DESC
            LIMIT ? OFFSET ?
        """

        paginated_params = query_params + [page_size, offset]
        cursor.execute(data_query, tuple(paginated_params))

        columns = [column[0] for column in cursor.description]
        transactions = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return {
            "data": transactions,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_items": total_items,
                "total_pages": total_pages,
            },
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred while fetching transaction reports: {e}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/transaction/export")
async def export_transactions_to_excel(
    date: str = Query(
        datetime.date.today().isoformat(),
        description="Date for the export (YYYY-MM-DD)",
    ),
    employee_group: Optional[str] = Query(
        None,
        description="Filter by employee group (case-insensitive, partial match)",
    ),
):
    conn = get_db_connection()
    try:
        day_start, day_end = _get_local_day_bounds_from_string(date)
        employee_group = employee_group.split()
        cursor = conn.cursor()

        query_transactions = """
            SELECT id, card_number, employee_id, employee_name, employee_group, tenant_id, tenant_name, transaction_date
            FROM transactions
            WHERE transaction_date >= ? AND transaction_date < ?
        """
        params_transactions = [day_start, day_end]
        if employee_group:
            query_transactions += " AND "
            for group in employee_group:
                query_transactions += " employee_group LIKE ? OR"
                params_transactions.append(f"%{group}%")    
            query_transactions = query_transactions[:-2]

        query_transactions += " ORDER BY transaction_date DESC"
        
        cursor.execute(query_transactions, tuple(params_transactions))

        transactions = [dict(row) for row in cursor.fetchall()]

        query_employee = "SELECT employee_id, name, employee_group FROM employees"
        params_employee = []
        if employee_group:
            query_employee += " WHERE "
            for group in employee_group:
                query_employee += " employee_group LIKE ? OR"
                params_employee.append(f"%{group}%")
            query_employee = query_employee[:-2]
            
        query_employee += " ORDER BY name"
        
        cursor.execute(query_employee, tuple(params_employee))
        
        employees = [dict(row) for row in cursor.fetchall()]

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Transactions"

        header_row_num = 3
        headers = [
            "No",
            "Employee ID",
            "Employee Name",
            "Employee Group",
            "Tenant Name",
            "Eat",
            "Transaction Date",
        ]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_num, value=header_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row = 4
        no = 1
        for employee in employees:
            transaction = next(
                (
                    t
                    for t in transactions
                    if t["employee_id"] == employee["employee_id"]
                ),
                None,
            )

            sheet.cell(row=current_row, column=1, value=no)
            sheet.cell(row=current_row, column=2, value=employee["employee_id"])
            sheet.cell(row=current_row, column=3, value=employee["name"])
            sheet.cell(row=current_row, column=4, value=employee["employee_group"])

            formatted_date = datetime.datetime.strptime(date, "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            )
            if transaction:
                sheet.cell(row=current_row, column=5, value=transaction["tenant_name"])
                sheet.cell(row=current_row, column=6, value=1)
                sheet.cell(row=current_row, column=7, value=formatted_date)
            else:
                sheet.cell(row=current_row, column=5, value="")
                sheet.cell(row=current_row, column=6, value=0)
                sheet.cell(row=current_row, column=7, value=formatted_date)

            no += 1
            current_row += 1

        employee_ids_set = {e["employee_id"] for e in employees}
        orphan_transactions = [
            t for t in transactions if t["employee_id"] not in employee_ids_set
        ]

        red_fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )

        for transaction in orphan_transactions:
            sheet.cell(row=current_row, column=1, value=no)
            sheet.cell(row=current_row, column=2, value=transaction["employee_id"])
            sheet.cell(
                row=current_row, column=3, value=transaction.get("employee_name", "N/A")
            )
            sheet.cell(
                row=current_row,
                column=4,
                value=transaction.get("employee_group", "N/A"),
            )
            sheet.cell(row=current_row, column=5, value=transaction["tenant_name"])
            sheet.cell(row=current_row, column=6, value=1)
            sheet.cell(row=current_row, column=7, value=date)

            for col in range(1, 8):
                sheet.cell(row=current_row, column=col).fill = red_fill
            no += 1
            current_row += 1

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for r in sheet.iter_rows(
            min_row=header_row_num, max_row=current_row - 1, min_col=1, max_col=7
        ):
            for cell in r:
                cell.border = thin_border

        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 33.57
        sheet.column_dimensions["D"].width = 15.28
        sheet.column_dimensions["E"].width = 12.14
        sheet.column_dimensions["F"].width = 5
        sheet.column_dimensions["G"].width = 15

        for col_letter in ["A", "B", "D", "F", "G"]:
            for cell in sheet[col_letter]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)

        return Response(
            content=virtual_workbook.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=transactions-{formatted_date}.xlsx"
            },
        )
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred while exporting transaction reports: {e}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/transaction/export/monthly")
async def export_transactions_to_excel(
    date: str = Query(
        datetime.date.today(),
        description="Date for the export (YYYY-MM)",
    ),
    employee_group: Optional[str] = Query(
        None,
        description="Filter by employee group (case-insensitive, partial match)",
    ),
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        date = date.split("-")
        if len(date) >= 2:
            date = date[0] + "-" + date[1]

        employee_group = employee_group.split()

        query_transactions = (
            "SELECT * FROM transactions WHERE strftime('%Y-%m', transaction_date) = ?"
        )
        params_transactions = [date]
        if employee_group:
            query_transactions += " AND "
        for group in employee_group:
            query_transactions += " employee_group LIKE ? OR"
            params_transactions.append(f"%{group}%")
        if employee_group:
            query_transactions = query_transactions[:-2]

        cursor.execute(query_transactions, tuple(params_transactions))

        transactions = [dict(row) for row in cursor.fetchall()]

        query_employees = "SELECT * FROM employees"
        params_employees = []
        if employee_group:
            query_employees += " WHERE "
        for group in employee_group:
            query_employees += " employee_group LIKE ? OR"
            params_employees.append(f"%{group}%")
        if employee_group:
            query_employees = query_employees[:-2]
        query_employees += " ORDER BY name"
        cursor.execute(query_employees, tuple(params_employees))
        employees = [dict(row) for row in cursor.fetchall()]
        employee_data = {}
        for emp in employees:
            employee_data[emp["employee_id"]] = {
                "details": emp,
                "eats_by_day": defaultdict(str),
            }

        for t in transactions:
            emp_id = t["employee_id"]
            if emp_id in employee_data:
                day = datetime.datetime.fromisoformat(t["transaction_date"]).day
                employee_data[emp_id]["eats_by_day"][day] = t["tenant_name"]

        wb = Workbook()
        sheet = wb.active
        sheet.title = f"Transactions {date}"

        year, month = map(int, date.split("-"))
        _, num_days = calendar.monthrange(year, month)

        header_row_num = 3
        static_headers = ["No", "Employee ID", "Employee Name", "Employee Group"]
        date_headers = [str(day) for day in range(1, num_days + 1)]
        all_headers = static_headers + date_headers

        for col_num, header_title in enumerate(all_headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_num, value=header_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row = 4
        no = 1
        for emp in employees:
            emp_id = emp["employee_id"]
            data = employee_data[emp_id]

            sheet.cell(row=current_row, column=1, value=no)
            sheet.cell(row=current_row, column=2, value=data["details"]["employee_id"])
            sheet.cell(row=current_row, column=3, value=data["details"]["name"])
            sheet.cell(
                row=current_row, column=4, value=data["details"]["employee_group"]
            )

            for day in range(1, num_days + 1):
                eat_value = data["eats_by_day"].get(day, '')
                sheet.cell(
                    row=current_row, column=len(static_headers) + day, value=eat_value
                )

            current_row += 1
            no += 1

        employee_ids_set = {e["employee_id"] for e in employees}
        orphan_transactions = [
            t for t in transactions if t["employee_id"] not in employee_ids_set
        ]

        orphan_data = {}
        for t in orphan_transactions:
            emp_id = t["employee_id"]
            if emp_id not in orphan_data:
                orphan_data[emp_id] = {
                    "details": t,
                    "eats_by_day": defaultdict(str),
                }
            day = datetime.datetime.fromisoformat(t["transaction_date"]).day
            orphan_data[emp_id]["eats_by_day"][day] = t["tenant_name"]

        red_fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )

        for emp_id, data in orphan_data.items():

            sheet.cell(row=current_row, column=1, value=no)
            sheet.cell(row=current_row, column=2, value=emp_id)
            sheet.cell(
                row=current_row,
                column=3,
                value=data["details"].get("employee_name", "N/A"),
            )
            sheet.cell(
                row=current_row,
                column=4,
                value=data["details"].get("employee_group", "N/A"),
            )

            for day in range(1, num_days + 1):
                eat_value = data["eats_by_day"].get(day, 0)
                sheet.cell(
                    row=current_row, column=len(static_headers) + day, value=eat_value
                )

            for col in range(1, len(all_headers) + 1):
                sheet.cell(row=current_row, column=col).fill = red_fill

            current_row += 1
            no += 1

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for r in sheet.iter_rows(
            min_row=header_row_num,
            max_row=current_row - 1,
            min_col=1,
            max_col=len(all_headers),
        ):
            for cell in r:
                cell.border = thin_border

        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 33.57
        sheet.column_dimensions["D"].width = 15.28

        for i in range(len(static_headers) + 1, len(all_headers) + 1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = 4
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)

        return Response(
            content=virtual_workbook.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=transactions-{date}.xlsx"
            },
        )
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred while exporting transaction reports: {e}",
        )
    finally:
        if conn:
            conn.close()
            
